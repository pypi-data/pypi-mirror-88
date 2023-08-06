#!/usr/bin/env python
# -*- coding:utf-8 -*-

import csv
from pathlib import Path
from typing import Union
import openpyxl


class Recorder(object):
    def __init__(self, file_path: str = None, cache: int = 50):
        """初始化                                                    \n
        :param file_path: 保存的文件路径
        :param cache: 每接收多少个写入文件
        """
        self.cache = cache
        self.file_path = file_path
        self._data = []
        self._before = []
        self._after = []

    def __del__(self) -> None:
        """对象关闭时把剩下的数据写入文件"""
        if self._data:
            self.record()

    def add_data(self, data: Union[list, tuple]) -> None:
        """添加数据                                                           \n
        :param data: 插入的数据，元组或列表
        :return: None
        """
        self._data.extend(data)
        if len(self._data) >= self.cache:
            self.record()

    def record(self) -> None:
        """记录数据"""
        if self._data:
            if self.file_path.endswith('.xlsx'):
                self._to_excel()
            else:
                with open(self.file_path, 'a+', newline='') as f:
                    csv_write = csv.writer(f)
                    for i in self._data:
                        csv_write.writerow(self._before + i + self._after)

            self._data = []

    def _to_excel(self) -> None:
        """把数据记录到excel"""
        if Path(self.file_path).exists():
            wb = openpyxl.load_workbook(self.file_path)
        else:
            wb = openpyxl.Workbook()

        ws = wb.active
        for i in self._data:
            ws.append(self._before + i + self._after)

        wb.save(self.file_path)
        wb.close()

    def set_before(self, before: Union[list, tuple, str]) -> None:
        """在数据前面补充的列                                                    \n
        :param before: 列表、元组或字符串，为字符串时则补充一列
        :return: None
        """
        self.record()

        if not before:
            self._before = []
        elif isinstance(before, (list, tuple)):
            self._before = list(before)
        elif isinstance(before, str):
            self._before = [before]
        else:
            raise TypeError('只能接收list、tuple、str或空。')

    def set_after(self, after: Union[list, tuple, str]) -> None:
        """在数据后面补充的列                                                      \n
        :param after: 列表、元组或字符串，为字符串时则补充一列
        :return: None
        """
        self.record()

        if not after:
            self._after = []
        elif isinstance(after, (list, tuple)):
            self._after = list(after)
        elif isinstance(after, str):
            self._after = [after]
        else:
            raise TypeError('只能接收list、tuple、str或空。')

    def add_head(self, head: Union[list, tuple]) -> None:
        """设置表头                                                    \n
        :param head: 表头，列表或元组
        :return: None
        """
        if self.file_path.endswith('.xlsx'):
            self._set_excel_head(head)
        else:
            with open(self.file_path, 'w+', newline='') as f:
                content = f.read()
                f.seek(0, 0)
                text = ','.join(head)
                f.write(f'{text}\n{content}')

    def _set_excel_head(self, head: Union[list, tuple]) -> None:
        """设置excel表头                                                    \n
        :param head: 表头，列表或元组
        :return: None
        """
        if Path(self.file_path).exists():
            wb = openpyxl.load_workbook(self.file_path)
        else:
            wb = openpyxl.Workbook()

        ws = wb.active
        ws.insert_rows(1)

        for key, i in enumerate(head, 1):
            ws.cell(1, key).value = i

        wb.save(self.file_path)
        wb.close()
