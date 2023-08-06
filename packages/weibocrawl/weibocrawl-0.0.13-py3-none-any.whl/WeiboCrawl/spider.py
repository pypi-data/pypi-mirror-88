#coding=utf-8

import os
import time
import re
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import numbers, Alignment
from openpyxl import load_workbook
from pymongo import MongoClient
import pandas as pd
import numpy as np

#带入程序包
import json
import pandas as pd
import jieba
import jieba.posseg as psg
import collections
from wordcloud import WordCloud, ImageColorGenerator, STOPWORDS
import matplotlib.pyplot as plt
from cnsenti import Sentiment
from PIL import Image


rootPath=r'./'

class Excel():
    #初始化
    def __init__(self):
        self.path = ''
    # 创建或打开Excel
    def open(self, path, flag=False):
        self.path = path.encode('utf-8').decode('utf-8')
        if os.path.exists(self.path) and flag:
            self.wb=load_workbook(self.path)
        else:
            self.wb=Workbook()
        self.sheet_init('Sheet')
        # 初始化工作表
    def sheet_init(self, name):
        self.ws = self.wb[name] if name in self.wb.sheetnames else self.wb.create_sheet(name.encode('utf-8').decode('utf-8'))
        # 设置列宽
        # self.ws.column_dimensions.fitToWidth=False
        # self.ws.column_dimensions['A'].width = 10
        # self.ws.column_dimensions['B'].width = 30
        # self.ws.column_dimensions['C'].width = 28
        # self.ws.column_dimensions['D'].width = 30
        # self.ws.column_dimensions['L'].width = 3000

    # 按行写入Excel
    def write_row(self, row):
        self.ws.append(row)
    # 关闭Excel
    def close(self):
        self.wb.close()
    # 格式化输出
    def format(self):
        for row in self.ws.rows:
            for i in range(self.ws.min_column, self.ws.max_column):
                row[i].alignment = Alignment(horizontal='right')
                # row[i].number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1 if i<0 else '#,###'
    # 保存并关闭Excel
    def save(self):
        self.wb.save(filename=self.path)
        # self.wb.close()


class Mongo():
    #初始化
    def __init__(self,stamp=None):
        self.excel = Excel()  # Excel对象
        self.stamp = (datetime.now()).strftime('%Y%m%d') if stamp is None else stamp
        self.client = MongoClient("localhost", 27017)
        self.db=self.client['baidunews']

    def getStamp(self):
        return self.utc2local(self.local2utc(datetime.now())).strftime('%Y%m%d')

    # 下载数据并保存
    def export(self):
        # 打开Excel
        # # 大数据，区块链，其他，人工智能，网络金融，云计算
        self.excel.open(rootPath + 'ExcelData/search_汇总1.xlsx')
        title='company,class, keyword,title,url,summary,source,stamp'.split(',')
        self.excel.write_row(title)
        # 写入数据
        self._matches = self.db['search_汇总1']
        # self._matches = self.db['guyulab_2']
        # self._matches = self.db['zhenshigushi1_2']
        try:
            matches = self._matches.find()
        except:
            print (self.stamp + " error, quit()")
            quit()
        for match in matches:
            try:
                # arctile=match['html']
                # arctile='error' if (arctile is None) or (len(arctile)<2000) else str(arctile)
                body=[match['company'],match['fenlei'],match['keyword'],match['title'],match['url'],match['summary'],match['source'],match['stamp']]
                self.excel.write_row(body)
            except:
                continue
        # 关闭并保存Excel
        # self.excel.format()
        self.excel.save()

    # 下载数据并保存
    def pb_export(self):
        # 大数据，区块链，其他，人工智能，网络金融，云计算
        self._matches = self.db['search_其他1']
        self._matches2 = self.db['search_汇总1']
        try:
            # 不为空查询
            matches = self._matches.find()
            # matches = self._matches.find({'datetime': {'$lte':1598528644,'$gte':1567424487}, 'content': {'$not': {'$in': [None]}}, 'author': {'$not': {'$in': ['']}}})
            # matches = self._matches.find(
            #     {'datetime': {'$lte': 1599015994, '$gte': 1567135988}, 'content': {'$in': [None]}})

            print(matches.count())

            for match in matches:
                try:
                    self._matches2.insert_one(match)
                    # company=match['company']
                    # # fenlei = match['fenlei']
                    # keyword = match['keyword']
                    # title = match['title']
                    # summary = match['summary']
                    #
                    # content=title+summary
                    # flag=(company in content) and (keyword in content)
                    # print(flag)
                    # if flag:
                    #     self._matches2.insert_one(match)

                except:
                    continue
        except:
            print (self.stamp + " error, quit()")
            quit()

    # 下载数据并保存
    def pb_export2(self):

        self._matches = self.db['renwumag1980_2']
        try:
            # 不为空查询
            # matches = self._matches.find()
            matches = self._matches.find({'video_num': {'$gte': 1}})
            # matches = self._matches.find({'img_num': {'$lte': 3}})
            # df = pd.DataFrame(matches)
            # # print(df)
            # arcwnum=df['key1']
            # print(df['key1'])
            #
            # fanwei = list(range(0, 16000, 2000))
            # fenzu = pd.cut(arcwnum.values, fanwei, right=False)  # 分组区间,长度91
            # # print(fenzu.codes)  # 标签
            # # print(fenzu.categories)  # 分组区间，长度8
            # pinshu = fenzu.value_counts()  # series,区间-个数
            # print(pinshu)

            for match in matches:
                try:

                    img_num=match['img_num']
                    # chi = re.findall(r'[\u4E00-\u9FFF]', content)
                    # match['key1']=len(content)
                    print(img_num)
                    print(type(img_num))

                    # content=match['content']
                    # # chi = re.findall(r'[\u4E00-\u9FFF]', content)
                    # match['key1']=len(content)
                    # print(match['key1'])

                    # datetime=match['datetime']
                    # timeArray = time.localtime(datetime)
                    # match['key2']=timeArray[3]
                    # print(match['key2'])

                    # self._matches.update_one({'symbol': match['symbol']}, {"$set": match}, True)
                except:
                    continue
            print(matches.count())
        except:
            print (self.stamp + " error, quit()")
            quit()

    # 下载数据并保存
    def pb_export3(self):

        self._matches = self.db['guyulab_2']
        # self._matches = self.db['zhenshigushi1_2']
        # self._matches = self.db['renwumag1980_2']
        try:
            # 不为空查询
            matches = self._matches.find()
            # matches = self._matches.find({'author': {'$not': {'$in': ['']}}})
            df = pd.DataFrame(matches)
            # print(df)
            arcwnum=df['key2']
            # print(df['key1'])

            fanwei = list(range(0, 25, 1))
            # fanwei = list(range(0, 18000, 2000))
            fenzu = pd.cut(arcwnum.values, fanwei, right=False)  # 分组区间,长度91
            # print(fenzu.codes)  # 标签
            # print(fenzu.categories)  # 分组区间，长度8
            pinshu = fenzu.value_counts()  # series,区间-个数
            print(pinshu)

            # pinshu.plot(kind='bar')
            # plt.text(0,29,str(29))

            qujian = pd.cut(arcwnum, fanwei, right=False)
            df['区间'] = qujian.values
            df.groupby('区间').median()
            df.groupby('区间').mean()  # 每个区间平均数

            pinshu_df = pd.DataFrame(pinshu, columns=['频数'])
            pinshu_df['频率f'] = pinshu_df / pinshu_df['频数'].sum()
            pinshu_df['频率%'] = pinshu_df['频率f'].map(lambda x: '%.2f%%' % (x * 100))

            pinshu_df['累计频率f'] = pinshu_df['频率f'].cumsum()
            pinshu_df['累计频率%'] = pinshu_df['累计频率f'].map(lambda x: '%.4f%%' % (x * 100))

            print(pinshu_df)
            pinshu_df.to_csv('人物.csv')
            # xticks=[x for x in range(0, 24, 1)]
            # print(xticks)

            # pinshu_df.plot(kind='line', marker='o', color='g', title='发布时间')

            pinshu_df.plot.bar()
            # Temperature = pd.DataFrame(pinshu_df['频数'],
            #                            index=fanwei)
            # Temperature.plot(kind='line', marker='o', color='r', title='温度变化')
            # fig = plt.figure()
            # fig.s
            # ax = fig.add_subplot(1, 1, 1)
            # plt.set_xlabel('Stages')
            # plt.plot(pinshu_df)
            plt.show()
        except:
            print (self.stamp + " error, quit()")
            quit()


    # 下载数据并保存
    def pb_export4(self):
        self._matches = self.db['guyulab_2']
        # self._matches = self.db['zhenshigushi1_2']
        # self._matches = self.db['renwumag1980_2']
        try:
            # 不为空查询
            matches = self._matches.find()
            # matches = self._matches.find({'video_num': {'$gte': 1}})
            # matches = self._matches.find({'img_num': {'$lte': 3}})


            for match in matches[0:1]:
                try:
                    content=match['content']
                    title = match['title']
                    print(content)

                    # 文本预处理
                    pattern = re.compile(u'\t|\n|\.|-|:|;|\)|\(|\?|"')
                    data = re.sub(pattern, '', content)

                    print([(x.word, x.flag) for x in psg.cut(data)])      #词性分析

                    # 文本分词--精确模式分词
                    seg_list_exact = jieba.cut(data, cut_all=False)

                    object_list = []
                    # 自定义常见去除词库
                    remove_words = [u'的', u'，', u'和', u'是', u'随着', u'对于', u'对', u'等', u'能', u'都', u'。', u' ', u'、',
                                    u'中', u'在', u'了', u'通常', u'如果', u'我们', u'需要']
                    for word in seg_list_exact:
                        if word not in remove_words:
                            object_list.append(word)

                    # 去除单个词
                    for i in range(len(object_list) - 1, -1, -1):
                        if (len(object_list[i]) < 2):
                            object_list.pop(i)

                    # 对分词做词频统计
                    word_counts = collections.Counter(object_list)
                    # 获取前100最高频的词
                    word_counts_top100 = word_counts.most_common(100)

                    data=json.dumps(word_counts_top100, ensure_ascii=False)
                    print(word_counts_top100)

                    match['word_top100']=data

                    # # 绘制词云
                    # my_wordcloud = WordCloud(
                    #     background_color='white',  # 设置背景颜色
                    #     # mask=img,  # 背景图片
                    #     max_words=200,  # 设置最大显示的词数
                    #     stopwords=STOPWORDS,  # 设置停用词
                    #     # 设置字体格式，字体格式 .ttf文件需自己网上下载，最好将名字改为英文，中文名路径加载可能会出现问题。
                    #     font_path='simhei.ttf',
                    #     max_font_size=100,  # 设置字体最大值
                    #     random_state=50,  # 设置随机生成状态，即多少种配色方案
                    #     ##提高清晰度
                    #     width=1000, height=600,
                    #     min_font_size=20,
                    # ).generate_from_frequencies(word_counts)

                    # 显示生成的词云图片
                    # plt.imshow(my_wordcloud)
                    # plt.axis('off')
                    # plt.show()
                    # my_wordcloud.to_file("./ImgData/"+title+".png")

                    self._matches.update_one({'symbol': match['symbol']}, {"$set": match}, True)
                except:
                    continue
            print(matches.count())
        except:
            print (self.stamp + " error, quit()")
            quit()


    # 下载数据并保存
    def pb_export5(self):
        # self.excel.open(rootPath + 'ExcelData/m' + self.stamp + '.xlsx')
        # self._matches = self.db['guyulab_2']
        # self._matches = self.db['zhenshigushi1_2']
        self._matches = self.db['renwumag1980_2']

        taidu=''
        count=0

        senti = Sentiment()

        try:
            # 不为空查询
            matches = self._matches.find()
            # matches = self._matches.find({'sentiment': '负面'})
            # matches = self._matches.find({'img_num': {'$lte': 3}})

            key3_dic={}
            for match in matches[0:1]:
                try:
                    pass
                    # key3=match['author'] if match['author'] is not None else '转载其他'
                    # if key3 in key3_dic.keys():
                    #     key3_dic[key3]+=1
                    # else:
                    #     key3_dic[key3] = 1

                    content = match['content']
                    print(content)


                    # test_text = '我好开心啊，非常非常非常高兴！今天我得了一百分，我很兴奋开心，愉快，开心'
                    result = senti.sentiment_count(content)
                    print(senti.Negs)
                    taidu='正面' if (result['pos']>result['neg']) else '负面'
                    if (abs(result['pos']-result['neg']))<10:
                        count+=1

                    # match['sentiment']=taidu
                    # match['like_num']=result['pos']
                    # match['old_like_num'] = result['neg']
                    #
                    #
                    # self._matches.update_one({'symbol': match['symbol']}, {"$set": match}, True)
                except:
                    continue
            # print(key3_dic)
            # self.excel.write_row([x for x in key3_dic.keys()])
            # self.excel.write_row([x for x in key3_dic.values()])
            # self.excel.save()

            print(matches.count())
        except:
            print (self.stamp + " error, quit()")
            quit()

# 主程序入口
if __name__ == '__main__':
    # Ctrl+ / 切换注释
    # db = Mongo("2019-04-03")   #下载历史数据（但必须已写入数据库）

    # 显示所有列
    # pd.set_option('display.max_columns', None)

    # plt.rcParams['font.sans-serif'] = ['SimHei']  # 显示中文标签
    # plt.rcParams['axes.unicode_minus'] = False


    # # 显示所有行
    # pd.set_option('display.max_rows', None)
    # # 设置value的显示长度为100，默认为50
    # pd.set_option('max_colwidth', 100)

    db = Mongo()
    # db.pb_export()
    db.export()


    print('\nTS:'+datetime.now().strftime('%Y-%m-%d %H:%M:%S'))



