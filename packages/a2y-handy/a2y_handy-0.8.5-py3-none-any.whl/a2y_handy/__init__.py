from typing import List as _List, Callable as _Callable, Tuple as _Tuple, Union as _Union
from threading import Lock as _Lock
from datetime import datetime as _datetime
from time import time as _now
import sys as _sys
import traceback as _traceback
from openpyxl import load_workbook as _load_xlsx
from typing import List as _List, Dict as _Dict


def int_2_bool_list(integer, bit_count) -> _List[bool]:
	result = [False] * bit_count
	for i in range(bit_count):
		if (integer & (1 << i)) != 0:
			result[i] = True
	return result


def bool_list_2_int(bools: _List[bool]) -> int:
	result = 0
	for b, idx in enumerate(bools):
		if b:
			result |= (1 << idx)

	return result


class Callback:
	def __init__(self):
		self._callback_list = list()
		self._callback_list_guard = _Lock()

	def subscribe(self, callback: _Callable):
		with self._callback_list_guard:
			if callback not in self._callback_list:
				self._callback_list.append(callback)

	def unsubscribe(self, callback: _Callable):
		with self._callback_list_guard:
			if callback in self._callback_list:
				self._callback_list.remove(callback)

	def unsubscribe_all(self):
		with self._callback_list_guard:
			self._callback_list.clear()

	def __call__(self, *args, **kwargs):
		with self._callback_list_guard:
			for callback in self._callback_list:
				callback(*args, **kwargs)


def timestamp(stamp: float = None) -> str:
	if stamp is None:
		stamp = _now()
	dt = _datetime.fromtimestamp(stamp)
	return dt.isoformat(' ')[:19]


def timestamp_as_filename(stamp: float = None, sep: str = '-') -> str:
	return timestamp(stamp).replace(':', sep)


def format_exception(skip_level: int = 0):
	exc_type, exc_value, exc_tb = _sys.exc_info()
	for i in range(skip_level):
		if exc_tb.tb_next is not None:
			exc_tb = exc_tb.tb_next

	return _traceback.format_exception(exc_type, exc_value, exc_tb)


def print_exception(skip_level: int = 0):
	lines = format_exception(skip_level)
	print(''.join(lines), end='')


class StrState:
	__states__: _Tuple[str] = ()

	@classmethod
	def all(cls) -> _Tuple[str]:
		return tuple(cls.__states__)

	def __init__(self, value: _Union[str]):
		if isinstance(value, str):
			assert value in self.__states__
			self.__value = value
		else:
			assert isinstance(value, StrState)
			self.__value = value.__value

	def set(self, value: _Union[str]):
		if isinstance(value, str):
			assert value in self.__states__
			self.__value = value
		else:
			assert isinstance(value, StrState)
			self.__value = value.__value

	def __contains__(self, item):
		if isinstance(item, str):
			return item in self.__states__
		else:
			assert isinstance(item, StrState)
			return item.__value in self.__states__

	def __str__(self):
		return self.__value

	def __eq__(self, other):
		if isinstance(other, str):
			assert other in self.__states__
			return self.__value == other
		assert isinstance(other, StrState)
		return self.__value == other.__value

	def __int__(self):
		return self.__states__.index(self.__value)

	def __lt__(self, other):
		if isinstance(other, StrState):
			return int(self) < int(other)
		else:
			assert isinstance(other, str)
			return int(self) < self.__states__.index(other)


def load_a_list_of_dict_from_xlsx(
		xlsx_name: str,
		sheet: _Union[int, str] = 0,
		column_count: int = 0,
		row_count: int = 0
) -> _List[dict]:
	book = _load_xlsx(xlsx_name, read_only=True, data_only=True)
	if isinstance(sheet, int):
		page = book.worksheets[sheet]
	else:
		page = book.get_sheet_by_name(sheet)
	headers = []
	column_index = 1
	while column_index <= column_count or column_count <= 0:
		header = page.cell(1, column_index).value
		if header is None or header == '':
			column_count = column_index - 1
			break
		else:
			headers.append(header)
			column_index += 1
	data_rows = []
	row_index = 2
	while row_count == 0 or (row_index - 1) <= row_count:
		data_row = []
		if page.cell(row_index, 1).value in [None, '']:
			break
		for _column_idx in range(column_count):
			column_index = _column_idx + 1
			value = page.cell(row_index, column_index).value
			data_row.append(value)
		data_rows.append(data_row)
		row_index += 1

	result = []
	for data_row in data_rows:
		item = dict()
		for index in range(column_count):
			item[headers[index]] = data_row[index]
		result.append(item)

	return result


def load_a_dict_of_dict_from_xlsx(
		xlsx_name: str,
		sheet: _Union[int, str] = 0,
		column_count: int = 0,
		row_count: int = 0,
		key: _Union[int, str] = 0
) -> _Dict[str, dict]:
	book = _load_xlsx(xlsx_name, read_only=True, data_only=True)
	if isinstance(sheet, int):
		page = book.worksheets[sheet]
	else:
		page = book.get_sheet_by_name(sheet)
	headers = []
	column_index = 1
	while column_index <= column_count or column_count <= 0:
		header = page.cell(1, column_index).value
		if header is None or header == '':
			column_count = column_index - 1
			break
		else:
			headers.append(header)
			column_index += 1

	if isinstance(key, int):
		key_index = key
	elif isinstance(key, str):
		key_index = headers.index(key)
	else:
		raise TypeError('Argument "key" must be an object of int or str.')

	data_rows = []
	row_index = 2
	while row_count == 0 or (row_index - 1) <= row_count:
		data_row = []
		if page.cell(row_index, 1).value in [None, '']:
			break
		for _column_idx in range(column_count):
			column_index = _column_idx + 1
			value = page.cell(row_index, column_index).value
			data_row.append(value)
		data_rows.append(data_row)
		row_index += 1

	result = dict()
	for data_row in data_rows:
		item = dict()
		key_value = data_row[key_index]
		if key_value is None:
			raise ValueError('None key found.')
		for index in range(column_count):
			item[headers[index]] = data_row[index]
		result[key_value] = item

	return result
