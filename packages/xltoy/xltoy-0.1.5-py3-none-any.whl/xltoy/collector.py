from openpyxl import load_workbook
from openpyxl.worksheet.cell_range import CellRange
from openpyxl.cell import Cell
from collections import defaultdict
from os.path import isfile
from datetime import datetime
from . import log, version

import re
import yaml
import dictdiffer

def useless_range(rng):
    """
    Check if rng is valid and usefull range
    :param rng:
    :return: True | False
    """
    if not rng:
        return True



class Sheet:
    pass

class Collector:
    label_names = [re.compile('^varname', re.IGNORECASE),
                   re.compile('^label', re.IGNORECASE)]
    model_names = [re.compile('^model', re.IGNORECASE)]
    data_names = [re.compile('^data', re.IGNORECASE)]

    def __init__(self, url:str, only_data:bool=False, relative:bool=False, add_fingerprint:bool=False):
        """
        Main data injestion, we need 2 instances of the sheet:
          - wb_data: with static data
          - wb: with the formulas

        :param url: url to an input excel file
        :param only_data: read only static values
        :param relative: all area are treated as if they starts from Row1 Col1
        """
        if not isfile(url):
            raise FileNotFoundError("File {} does not exists".format(url))

        self.sheets = {}
        self.pseudo = {}
        self.url = url
        self.anon_models = {}

        self.labels_as_data = True
        self.models_as_data = only_data
        log.debug("Labels read as {}".format( 'data' if self.labels_as_data else 'formula'))
        log.debug("Models read as {}".format('data' if self.models_as_data else 'formula'))

        self.wb_data = load_workbook(filename=self.url, data_only=True)
        if only_data:
            self.wb = self.wb_data
        else:
            self.wb = load_workbook(filename=self.url)
        self.relative = relative
        self.add_fingerprint = add_fingerprint

        self.sheetnames = self.wb_data.sheetnames
        self.named_ranges = self.wb_data.get_named_ranges()
        self.set_ranges()

        self.models = self.handle_range(
                        self.model_names,
                        self.models_as_data)

        self.labels = self.handle_range(
                        self.label_names,
                        self.labels_as_data)

        self.data = self.handle_range(
                      self.data_names,
                      use_data=True)


        # Check for anonimous models
        for k,v in self.models.items():
            if k not in self.labels:
                anon_labels = {'anon_{}'.format(x):'anon_{}'.format(x) for x in range(1,len(v))}
                self.anon_models[k] = '{} anon labels assigned'.format(len(anon_labels))
                self.labels[k] = anon_labels
        self.set_pseudo_excel()
        for k,v in self.anon_models.items():
            log.info("Found anonimous model {} : {}".format(k,v))


    def set_ranges(self):
        """
        Iter over named ranges and collect ALL ranges into:
            - self.ranges
            - self.params
            - self.text_ranges

        here we dont know if ranges are used, but we cannot distinguish cell_alias, and other stuff.

        :return:
            None
        """
        self.ranges = {}
        self.text_ranges = {}
        self.params = {}

        for x in self.named_ranges:
            for sheet, rng in x.destinations:
                if x.type == 'TEXT':
                    self.text_ranges[x.name] = x.value
                elif x.type == 'NUMBER':
                    self.params[x.name] = x.value
                else:
                    if useless_range(rng):
                        log.debug("{} range [{}]in sheet {} discarded ".format(x.name,rng,sheet))
                        continue
                    try:
                        cells = self.wb_data[sheet][rng]
                    except KeyError as err:
                        log.error(err)
                        continue

                    if isinstance(cells,Cell):
                        # Single named cell
                        self.ranges[x.name] = (cells,)
                    else:
                        # Range named
                        self.ranges[x.name] = sum(cells, ())

        log.debug("{} named ranges collected".format(len(self.ranges)))
        log.debug("{} parameters collected".format(len(self.params)))
        log.debug("{} text range collected".format(len(self.text_ranges)))

    def to_relative(self, rng_name:str, cell:Cell) -> str:
        """
        :param rng_name:
        :param cell:
        :return:
        """
        ref_cell = self.ranges[rng_name][0]
        min_row, min_col = ref_cell.row, ref_cell.column
        return "R{0}C{0}".format(cell.row - min_row + 1, cell.column - min_col + 1)


    def handle_range(self, labels: list, use_data:bool):
        """
        iter over collected ranges and
        :param labels:
        :param use_data:
        :return:
        """
        coll = {}
        for lbl in self.ranges:
            for check_valid in labels:
                if check_valid.match(lbl):
                    cells = self.ranges[lbl]
                    sheet_names = set([x.parent.title for x in cells])
                    if len(sheet_names) > 1:
                        raise NotImplementedError("Range defined on multiple sheets is not handled")
                    sheet_name = sheet_names.pop()
                    if use_data:
                        coll[sheet_name] = {self.to_relative(lbl,x) if self.relative else x.coordinate:x.value
                                            for x in cells}
                    else:
                        coll[sheet_name] = {self.to_relative(lbl,x) if self.relative else x.coordinate:self.wb[sheet_name][x.coordinate].value
                                            for x in cells}
                    break
        return coll

    def set_pseudo_excel(self):
        if not self.pseudo:
            # Remember dict are ordered!
            if self.add_fingerprint:
                self.pseudo['xltoy.version'] = version
                self.pseudo['xltoy.filename'] = self.url
                self.pseudo['xltoy.datetime'] = datetime.now().isoformat()

            for sheet, labels in self.labels.items():
                self.pseudo[sheet] = dict(zip(labels.values(), self.models[sheet].values()))
                if sheet in self.data:
                    self.pseudo[sheet]['data'] = self.data[sheet]

            # We must handle data ranges not touched before, here
            # we have only data range not in sheet with labels
            for sheet in set(self.data) - set(self.labels):
                self.pseudo[sheet] = self.data[sheet]

    def to_yaml(self):
        self.set_pseudo_excel()
        return yaml.dump(self.pseudo)


class YamlCollector(Collector):
    """
    Collector for yaml format, it is used only to set
    self.pseudo attribute.
    """
    def __init__(self, url):
        if not isfile(url):
            raise FileNotFoundError("File {} does not exists".format(url))

        with open(url) as fin:
            self.pseudo= yaml.load(fin, Loader=yaml.FullLoader)



class DiffCollector:
    def __init__(self, url1, url2, only_data:bool=False, relative:bool=False, add_fingerprint:bool=False):
        """
        workbook differ, given 2 files (excel, yaml or json) it can do intelligent comparison

        :param url1:
        :param url2:
        :param only_data: ignore formulas and compare only values
        :param relative: all area are treated as if they starts from Row1 Col1
        :param add_fingerprint:
        """
        c1,c2 = [YamlCollector(url) if url.lower().endswith('yaml') else Collector(url,only_data=only_data, relative=relative, add_fingerprint=add_fingerprint)
                 for url in [url1,url2]]

        self.iter_differs = dictdiffer.diff(c1.pseudo,c2.pseudo)
        self.diff = {}

    def to_yaml(self):
        if not self.diff:
            for kind, mid, sh_cells in self.iter_differs:
                if kind not in self.diff:
                    self.diff[kind] = {}
                if isinstance(mid, list):
                    if len(mid)==1:
                        mid = mid[0]
                    else:
                        raise RuntimeError("{} not understood".format(mid))
                if kind == 'change':
                    sheet, label = mid.split('.')
                    if sheet not in self.diff[kind]:
                        self.diff[kind][sheet] = {}
                    self.diff[kind][sheet][label] = ' -> '.join(['{}'.format(x) for x in sh_cells])
                else:
                    for sheet, cells in sh_cells:
                        if sheet not in self.diff[kind]:
                            self.diff[kind][sheet] = {}
                        self.diff[kind][sheet] = cells
        if self.diff:
            print(yaml.dump(self.diff))

