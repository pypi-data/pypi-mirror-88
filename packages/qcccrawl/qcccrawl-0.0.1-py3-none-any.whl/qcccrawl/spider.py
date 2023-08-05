#coding=utf-8

import os
import time
import re
import xlsxwriter
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import numbers, Alignment
from openpyxl import load_workbook
from pymongo import MongoClient
import pandas as pd
import numpy as np

#带入程序包
import json
import pandas as pd
import jieba
import jieba.posseg as psg
import collections
from wordcloud import WordCloud, ImageColorGenerator, STOPWORDS
import matplotlib.pyplot as plt
from cnsenti import Sentiment
from PIL import Image


rootPath=r'./'

class Excel():
    #初始化
    def __init__(self):
        self.path = ''
    # 创建或打开Excel
    def open(self, path, flag=False):
        self.path = path.encode('utf-8').decode('utf-8')
        if os.path.exists(self.path) and flag:
            self.wb=load_workbook(self.path)
        else:
            self.wb=Workbook()
        self.sheet_init('Sheet')
        # 初始化工作表
    def sheet_init(self, name):
        self.ws = self.wb[name] if name in self.wb.sheetnames else self.wb.create_sheet(name.encode('utf-8').decode('utf-8'))
        # 设置列宽
        # self.ws.column_dimensions.fitToWidth=False
        # self.ws.column_dimensions['A'].width = 10
        # self.ws.column_dimensions['B'].width = 30
        # self.ws.column_dimensions['C'].width = 28
        # self.ws.column_dimensions['D'].width = 30
        # self.ws.column_dimensions['L'].width = 3000

    # 按行写入Excel
    def write_row(self, row):
        self.ws.append(row)
    # 关闭Excel
    def close(self):
        self.wb.close()
    # 格式化输出
    def format(self):
        for row in self.ws.rows:
            for i in range(self.ws.min_column, self.ws.max_column):
                row[i].alignment = Alignment(horizontal='right')
                # row[i].number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1 if i<0 else '#,###'
    # 保存并关闭Excel
    def save(self):
        self.wb.save(filename=self.path)
        # self.wb.close()


class Mongo():
    #初始化
    def __init__(self,stamp=None):
        self.excel = Excel()  # Excel对象
        self.stamp = (datetime.now()).strftime('%Y%m%d') if stamp is None else stamp
        self.client = MongoClient("localhost", 27017)
        self.db=self.client['nytimes']
        self.db2 = self.client['weibo2']

    def getStamp(self):
        return self.utc2local(self.local2utc(datetime.now())).strftime('%Y%m%d')

    # 下载数据并保存
    def export(self):
        # 打开Excel
        # # 大数据，区块链，其他，人工智能，网络金融，云计算
        self.excel.open(rootPath + 'ExcelData/search_汇总1.xlsx')
        title='company,class, keyword,title,url,summary,source,stamp'.split(',')
        self.excel.write_row(title)
        # 写入数据
        self._matches = self.db['search_汇总1']
        # self._matches = self.db['guyulab_2']
        # self._matches = self.db['zhenshigushi1_2']
        try:
            matches = self._matches.find()
        except:
            print (self.stamp + " error, quit()")
            quit()
        for match in matches:
            try:
                # arctile=match['html']
                # arctile='error' if (arctile is None) or (len(arctile)<2000) else str(arctile)
                body=[match['company'],match['fenlei'],match['keyword'],match['title'],match['url'],match['summary'],match['source'],match['stamp']]
                self.excel.write_row(body)
            except:
                continue
        # 关闭并保存Excel
        # self.excel.format()
        self.excel.save()

    # 下载数据并保存
    def pb_export(self):
        # 大数据，区块链，其他，人工智能，网络金融，云计算
        self._matches = self.db['user']
        self._matches2 = self.db2['user']
        try:
            # 不为空查询
            matches = self._matches.find()
            # matches = self._matches.find({'datetime': {'$lte':1598528644,'$gte':1567424487}, 'content': {'$not': {'$in': [None]}}, 'author': {'$not': {'$in': ['']}}})
            # matches = self._matches.find(
            #     {'datetime': {'$lte': 1599015994, '$gte': 1567135988}, 'content': {'$in': [None]}})

            print(matches.count())

            for match in matches:
                try:
                    self._matches2.insert_one(match)
                    # company=match['company']
                    # # fenlei = match['fenlei']
                    # keyword = match['keyword']
                    # title = match['title']
                    # summary = match['summary']
                    #
                    # content=title+summary
                    # flag=(company in content) and (keyword in content)
                    # print(flag)
                    # if flag:
                    #     self._matches2.insert_one(match)

                except:
                    continue
        except:
            print (self.stamp + " error, quit()")
            quit()

    # 下载数据并保存
    def pb_export2(self):

        self._matches = self.db['search_COVID-19']
        try:
            # 不为空查询
            matches = self._matches.find()
            # matches = self._matches.find({'video_num': {'$gte': 1}})
            # matches = self._matches.find({'img_num': {'$lte': 3}})
            # df = pd.DataFrame(matches)

            # matches1 = self._matches1.find()

            df = pd.DataFrame(list(matches))
            del df['_id']

            print(df)
            df.to_excel('./' + 'search_COVID-19.xlsx', engine='xlsxwriter')




            # print(matches.count())
        except:
            print (self.stamp + " error, quit()")
            quit()

    # 下载数据并保存
    def pb_export3(self):

        self._matches = self.db['user']
        try:
            # 不为空查询
            matches = self._matches.find()
            # matches = self._matches.find({'video_num': {'$gte': 1}})
            # matches = self._matches.find({'img_num': {'$lte': 3}})
            # df = pd.DataFrame(matches)

            for match in matches:
                screen_name=match['screen_name']

                self._matches1 = self.db2[screen_name+'_comments']
                self._matches2 = self.db2[screen_name + '_commentsUser']
                self._matches3 = self.db2[screen_name + '_retweets']
                self._matches4 = self.db2[screen_name + '_retweetsUser']

                # df = pd.DataFrame(list(matches1))
                # del df['_id']
                # print(df)
                # df.to_excel(screen_name+'.xlsx')
                # matches1 = self._matches1.find()
                # for match1 in matches1:
                #     symbol=match1['comment_id']
                #     print(symbol)
                #     # 更新数据库
                #     match2 = self._matches2.find_one({'symbol': symbol})  # 代码
                #     flag = (match2 is None)
                #     if flag:  # 逐条操作数据库比较耗时；暂时关闭
                #         match2 = {}
                #         match2['symbol'] = match1['comment_id']
                #         match2['user_id'] = match1['user_id']
                #         match2['screen_name'] = match1['screen_name']
                #         match2['comment_id'] = match1['comment_id']
                #         match2['comment_name'] = match1['comment_name']
                #
                #     if flag:
                #         self._matches2.insert_one(match2)
                #     else:
                #         self._matches2.update_one({'symbol': match2['symbol']}, {"$set": match2}, True)

                matches3 = self._matches3.find()
                for match3 in matches3:
                    symbol=match3['retweet_id']
                    print(symbol)
                    # 更新数据库
                    match4 = self._matches4.find_one({'symbol': symbol})  # 代码
                    flag = (match4 is None)
                    if flag:  # 逐条操作数据库比较耗时；暂时关闭
                        match4 = {}
                        match4['symbol'] = match3['retweet_id']
                        match4['user_id'] = match3['user_id']
                        match4['screen_name'] = match3['screen_name']
                        match4['retweet_id'] = match3['retweet_id']
                        match4['retweet_name'] = match3['retweet_name']

                    if flag:
                        self._matches4.insert_one(match4)
                    else:
                        self._matches4.update_one({'symbol': match4['symbol']}, {"$set": match4}, True)

            # print(matches.count())
        except:
            print (self.stamp + " error, quit()")
            quit()

        # 下载数据并保存

    def pb_export4(self):

        self._matches = self.db['user']
        try:
            # 不为空查询
            matches = self._matches.find()
            # matches = self._matches.find({'video_num': {'$gte': 1}})
            # matches = self._matches.find({'img_num': {'$lte': 3}})
            # df = pd.DataFrame(matches)

            for match in matches:
                screen_name = match['screen_name']

                self._matches1 = self.db2[screen_name + '_commentsUser']
                self._matches2 = self.db2['comments_user']
                self._matches3 = self.db2[screen_name + '_retweetsUser']
                self._matches4 = self.db2['retweets_user']

                # df = pd.DataFrame(list(matches1))
                # del df['_id']
                # print(df)
                # df.to_excel(screen_name+'.xlsx')
                matches1 = self._matches1.find()
                for match1 in matches1:
                    symbol=match1['comment_id']
                    print(symbol)
                    # 更新数据库
                    match2 = self._matches2.find_one({'symbol': symbol})  # 代码
                    flag = (match2 is None)
                    if flag:  # 逐条操作数据库比较耗时；暂时关闭
                        match2 = {}
                        match2['symbol'] = match1['comment_id']
                        match2['user_id'] = match1['user_id']
                        match2['screen_name'] = match1['screen_name']
                        match2['comment_id'] = match1['comment_id']
                        match2['comment_name'] = match1['comment_name']

                    if flag:
                        self._matches2.insert_one(match2)
                    else:
                        self._matches2.update_one({'symbol': match2['symbol']}, {"$set": match2}, True)

                matches3 = self._matches3.find()
                for match3 in matches3:
                    symbol = match3['retweet_id']
                    print(symbol)
                    # 更新数据库
                    match4 = self._matches4.find_one({'symbol': symbol})  # 代码
                    flag = (match4 is None)
                    if flag:  # 逐条操作数据库比较耗时；暂时关闭
                        match4 = {}
                        match4['symbol'] = match3['retweet_id']
                        match4['user_id'] = match3['user_id']
                        match4['screen_name'] = match3['screen_name']
                        match4['retweet_id'] = match3['retweet_id']
                        match4['retweet_name'] = match3['retweet_name']

                    if flag:
                        self._matches4.insert_one(match4)
                    else:
                        self._matches4.update_one({'symbol': match4['symbol']}, {"$set": match4}, True)

            # print(matches.count())
        except:
            print(self.stamp + " error, quit()")
            quit()

    # 下载数据并保存
    def pb_export5(self):

        try:
            table_name='retweets_user'
            self._matches1 = self.db2[table_name]
            matches1 = self._matches1.find()
            df = pd.DataFrame(list(matches1))
            del df['_id']
            print(df)
            df.to_excel('./ExcelData/' + table_name + '.xlsx')
        except:
            print (self.stamp + " error, quit()")
            quit()


# 主程序入口
if __name__ == '__main__':
    # Ctrl+ / 切换注释
    # db = Mongo("2019-04-03")   #下载历史数据（但必须已写入数据库）

    # 显示所有列
    # pd.set_option('display.max_columns', None)
    # # 显示所有行
    # pd.set_option('display.max_rows', None)
    # # 设置value的显示长度为100，默认为50
    # pd.set_option('max_colwidth', 100)
    # plt.rcParams['font.sans-serif'] = ['SimHei']  # 显示中文标签
    # plt.rcParams['axes.unicode_minus'] = False



    db = Mongo()
    db.pb_export2()
    # db.export()


    print('\nTS:'+datetime.now().strftime('%Y-%m-%d %H:%M:%S'))



